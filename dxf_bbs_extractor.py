"""
dxf_bbs_pipeline.py
===================
Refactored core pipeline from dxf_bbs_extractor.py.
All DXF extraction, parsing, and BBS generation logic — no CLI, no UI.

FIXES in this version:
- is_link properly tracked on BarLabel
- suggest_shape_code() fixed to check individual suffixes
- build_summary() correctly identifies links per mark
- Shape code 60 properly written for links in BBS
- Extraction report includes Is Link column
"""
from __future__ import annotations

import math
import os
import re
import tempfile
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import ezdxf
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# 0. OPTIONAL: sans282_shape_codes.py import
# ---------------------------------------------------------------------------
try:
    import sans282_shape_codes as shape_codes
except ImportError:
    shape_codes = None

# ---------------------------------------------------------------------------
# 1. CONFIGURATION
# ---------------------------------------------------------------------------
BAR_LABEL_PATTERN = re.compile(
    r"""^\s*
        (?P<count>\d+)\s*
        (?P<type>[A-Za-z]+)\s*
        (?P<diameter>\d+)\s*
        -\s*
        (?P<mark>[A-Za-z0-9]+)
        (?:\s*-\s*(?P<spacing>\d+))?
        \s*(?P<suffix>.*?)\s*$
    """,
    re.VERBOSE,
)

TEXT_BEARING_TYPES = ("DIMENSION", "TEXT", "MTEXT")
LINK_KEYWORDS = ("LINK", "STIRRUP", "TIE")
MIN_VALID_LENGTH_MM = 50

DOT_LAYER_ALIASES = ("DIMENSION", "DIMENSIONS")
BAR_LAYER_ALIASES = ("REINFORCEMENT", "REINF")
LEADER_LINE_LAYER_ALIASES = ("DIMENSION", "DIMENSIONS")

DOT_LABEL_TOL_MM = 5
BAR_MATCH_TOL_MM = 10
PROXIMITY_MATCH_MAX_MM = 4000
PROXIMITY_AMBIGUITY_GAP_MM = 200

LEADER_X_ALIGN_TOL_MM = 5
LEADER_BEND_TOL_MM = 5
LEADER_TIP_TOL_MM = 400
LEADER_MAX_HOPS = 4
LEADER_START_AMBIGUITY_GAP_MM = 50
LEADER_RAY_CORRIDOR_MM = 200
LEADER_RAY_AMBIGUITY_GAP_MM = 150

MIN_DOT_RADIUS_MM = 3
MAX_DOT_RADIUS_MM = 100

LENGTH_VARIANCE_FLAG_MM = 100

WEIGHT_FACTORS = {8: 0.395, 10: 0.616, 12: 0.888, 16: 1.579, 20: 2.466, 25: 3.854, 32: 6.313}
REVIEW_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
DXF_LENGTH_FILL = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
LINK_SHAPE_FILL = PatternFill("solid", start_color="CFE2F3", end_color="CFE2F3")
PROXIMITY_LENGTH_FILL = PatternFill("solid", start_color="FCE5CD", end_color="FCE5CD")
HEADER_FONT = Font(name="Arial", bold=True, size=9)
BODY_FONT = Font(name="Arial", size=9)
THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REF_TABLE_START_COL = 19

# ---------------------------------------------------------------------------
# 2. DATA STRUCTURES
# ---------------------------------------------------------------------------
@dataclass
class BarLabel:
    layer: str
    entity_type: str
    handle: str
    raw_text: str
    x: float
    y: float
    measured_length: Optional[float] = None
    defpoint: Optional[tuple] = None
    anchor_source: Optional[str] = None
    count: Optional[int] = None
    bar_type: Optional[str] = None
    diameter: Optional[int] = None
    mark: Optional[str] = None
    spacing: Optional[int] = None
    suffix: str = ""
    parsed_ok: bool = False
    is_link: bool = False  # <-- FIXED: added is_link field


# ---------------------------------------------------------------------------
# 3. HELPERS
# ---------------------------------------------------------------------------
def _layer_matches(entity_layer: str, aliases: tuple[str, ...]) -> bool:
    return entity_layer.strip().upper() in {a.upper() for a in aliases}


def clean_dxf_text(raw: str) -> str:
    if raw is None:
        return ""
    text = raw
    text = re.sub(r"\\P", " ", text)
    text = re.sub(r"\\[A-Za-z][^;]*;", "", text)
    text = text.replace("{", "").replace("}", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_bar_label(raw_text: str) -> dict:
    text = clean_dxf_text(raw_text)
    m = BAR_LABEL_PATTERN.match(text)
    if not m:
        return {"parsed_ok": False}
    gd = m.groupdict()
    suffix = gd["suffix"].strip()
    return {
        "parsed_ok": True,
        "count": int(gd["count"]),
        "bar_type": gd["type"].upper(),
        "diameter": int(gd["diameter"]),
        "mark": gd["mark"].upper(),
        "spacing": int(gd["spacing"]) if gd["spacing"] else None,
        "suffix": suffix,
        "is_link": any(k in suffix.upper() for k in LINK_KEYWORDS),  # <-- FIXED
    }


def suggest_shape_code(suffix: str) -> Optional[str]:
    """Check if a suffix indicates a link/stirrup. Returns the literal
    string "LINK" (a flag, NOT a shape code) or None. Note: is_link is
    tracked directly on BarLabel/parse_bar_label now, so this function is
    mostly a leftover helper -- don't use its return value as a numeric
    shape code anywhere."""
    if not suffix:
        return None
    if any(k in suffix.upper() for k in LINK_KEYWORDS):
        return "LINK"
    return None


def _dist(p1: tuple, p2: tuple) -> float:
    return math.hypot(p1[0] - p2[0], p1[1] - p2[1])


def _point_to_segment_dist(p: tuple, a: tuple, b: tuple) -> float:
    px, py = p
    ax, ay = a
    bx, by = b
    dx, dy = bx - ax, by - ay
    if dx == 0 and dy == 0:
        return _dist(p, a)
    t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / (dx * dx + dy * dy)))
    return _dist(p, (ax + t * dx, ay + t * dy))


def _axis_aligned_dist(p1: tuple, p2: tuple) -> float:
    return min(abs(p1[0] - p2[0]), abs(p1[1] - p2[1]))


# ---------------------------------------------------------------------------
# 4b. BAR-CHAIN GEOMETRY (for shape-code detection)
# ---------------------------------------------------------------------------
# This is the piece that was dropped when dxf_bbs_extractor.py got
# refactored into this pipeline module -- ported back in from the
# original version. It groups connected LINE segments on the bar layer
# into full chains (not just single lines), so bent bars can be measured
# and classified by their real geometry instead of guessed.

CHAIN_SNAP_TOL_MM = 2  # endpoints within this distance are treated as "the same point"


@dataclass
class BarChain:
    segments: list         # [(start_xy, end_xy, length), ...] in path order
    segment_lengths: list  # just the lengths, in path order (candidate A, B, C, ...)
    total_length: float    # sum of segment lengths (raw geometry, NOT the calculated BBS length)
    num_segments: int
    turn_angles: list      # degrees of direction change at each internal vertex


def _chain_snap(p: tuple, tol: float = CHAIN_SNAP_TOL_MM) -> tuple:
    return (round(p[0] / tol), round(p[1] / tol))


def _turn_angle(seg1: tuple, seg2: tuple) -> float:
    """Angle (degrees, 0-180) the direction changes between two ordered,
    connected segments. 0 = continues straight, 90 = right-angle bend."""
    v1 = (seg1[1][0] - seg1[0][0], seg1[1][1] - seg1[0][1])
    v2 = (seg2[1][0] - seg2[0][0], seg2[1][1] - seg2[0][1])
    a1 = math.atan2(v1[1], v1[0])
    a2 = math.atan2(v2[1], v2[0])
    diff = math.degrees(a2 - a1)
    diff = (diff + 180) % 360 - 180
    return abs(diff)


def _order_chain(segs: list) -> Optional[list]:
    """Order a set of connected (p1, p2) segments into a single continuous
    path, flipping each segment's direction as needed. Returns None if the
    segments don't form one simple open path (e.g. they form a closed
    loop) -- caller should not guess at an ordering in that case."""
    counts = Counter()
    for p1, p2 in segs:
        counts[_chain_snap(p1)] += 1
        counts[_chain_snap(p2)] += 1
    free_ends = [pt for pt, c in counts.items() if c == 1]
    if len(free_ends) != 2:
        return None  # not a simple open path

    remaining = list(segs)
    used = [False] * len(remaining)
    ordered = []
    current_pt = None
    for i, (p1, p2) in enumerate(remaining):
        if _chain_snap(p1) == free_ends[0]:
            ordered.append((p1, p2)); used[i] = True; current_pt = p2; break
        if _chain_snap(p2) == free_ends[0]:
            ordered.append((p2, p1)); used[i] = True; current_pt = p1; break

    for _ in range(len(remaining) - 1):
        found = False
        for i, (p1, p2) in enumerate(remaining):
            if used[i]:
                continue
            if _chain_snap(p1) == _chain_snap(current_pt):
                ordered.append((p1, p2)); used[i] = True; current_pt = p2; found = True; break
            if _chain_snap(p2) == _chain_snap(current_pt):
                ordered.append((p2, p1)); used[i] = True; current_pt = p1; found = True; break
        if not found:
            return None
    return ordered


def build_bar_chains(doc, layer_aliases: tuple[str, ...] = BAR_LAYER_ALIASES) -> list:
    """Group connected LINE entities on the bar layer(s) into BarChains. A
    LINE that doesn't touch any other LINE becomes its own 1-segment chain."""
    msp = doc.modelspace()
    raw_lines = []
    for l in msp.query("LINE"):
        if not _layer_matches(l.dxf.layer, layer_aliases):
            continue
        s = (float(l.dxf.start[0]), float(l.dxf.start[1]))
        e = (float(l.dxf.end[0]), float(l.dxf.end[1]))
        raw_lines.append((s, e))

    point_to_lines: dict = defaultdict(list)
    for idx, (s, e) in enumerate(raw_lines):
        point_to_lines[_chain_snap(s)].append(idx)
        point_to_lines[_chain_snap(e)].append(idx)

    parent = list(range(len(raw_lines)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for pt, idxs in point_to_lines.items():
        unique_idxs = set(idxs)
        if len(unique_idxs) == 2:  # exactly two lines meet -> safe to chain
            a, b = list(unique_idxs)
            union(a, b)
        # len > 2 -> junction/branch: deliberately do NOT chain through it

    groups: dict = defaultdict(list)
    for idx in range(len(raw_lines)):
        groups[find(idx)].append(idx)

    chains = []
    for idxs in groups.values():
        segs = [raw_lines[i] for i in idxs]
        if len(segs) == 1:
            s, e = segs[0]
            length = _dist(s, e)
            chains.append(BarChain(segments=[(s, e, length)], segment_lengths=[round(length, 1)],
                                    total_length=length, num_segments=1, turn_angles=[]))
            continue

        ordered = _order_chain(segs)
        if ordered is None:
            # Couldn't resolve a clean single path (e.g. a closed loop) --
            # fall back to standalone segments rather than guess an order.
            for s, e in segs:
                length = _dist(s, e)
                chains.append(BarChain(segments=[(s, e, length)], segment_lengths=[round(length, 1)],
                                        total_length=length, num_segments=1, turn_angles=[]))
            continue

        seg_tuples = [(s, e, _dist(s, e)) for s, e in ordered]
        angles = [round(_turn_angle(ordered[i], ordered[i + 1]), 1) for i in range(len(ordered) - 1)]
        chains.append(BarChain(
            segments=seg_tuples,
            segment_lengths=[round(t[2], 1) for t in seg_tuples],
            total_length=sum(t[2] for t in seg_tuples),
            num_segments=len(seg_tuples),
            turn_angles=angles,
        ))
    return chains


def _chain_point_dist(dot: tuple, chain: BarChain) -> float:
    """Shortest distance from a point to any segment in the chain."""
    return min(_point_to_segment_dist(dot, s, e) for s, e, _ in chain.segments)


def suggest_shape_from_geometry(num_segments: int, turn_angles: list) -> tuple:
    """Ask sans282_shape_codes for shape-code candidates matching this
    geometry. Returns (auto_code, candidates) where auto_code is only set
    if EXACTLY ONE confirmed candidate exists -- otherwise None, since
    picking among several plausible confirmed codes (or accepting an
    unverified one) would be guessing, not detecting.

    A single segment is unambiguously shape 20 (straight bar) -- that's
    the library's own definition, not a guess, so it's always returned
    directly rather than run through the coarser multi-candidate matcher.
    """
    if num_segments == 1:
        return "20", ["20"]
    if shape_codes is None:
        return None, []
    try:
        candidates = shape_codes.match_shape_by_geometry(num_segments, turn_angles)
    except Exception:
        return None, []
    confirmed = [c for c in candidates if shape_codes.SHAPE_CODES[c].confidence == "confirmed" and c != "99"]
    auto_code = confirmed[0] if len(confirmed) == 1 else None
    return auto_code, candidates


# ---------------------------------------------------------------------------
# 4. DXF EXTRACTION
# ---------------------------------------------------------------------------
def get_entity_text(entity) -> Optional[str]:
    etype = entity.dxftype()
    if etype == "DIMENSION":
        override = entity.dxf.get("text", None)
        if override in (None, "", " "):
            return None
        return override
    if etype == "TEXT":
        return entity.dxf.text
    if etype == "MTEXT":
        try:
            return entity.plain_text()
        except Exception:
            return entity.text
    return None


def get_entity_position(entity) -> tuple[float, float]:
    etype = entity.dxftype()
    try:
        if etype == "DIMENSION":
            p = entity.dxf.text_midpoint if entity.dxf.hasattr("text_midpoint") else entity.dxf.defpoint
        elif etype == "TEXT":
            p = entity.dxf.insert
        elif etype == "MTEXT":
            p = entity.dxf.insert
        else:
            return (0.0, 0.0)
        return (float(p[0]), float(p[1]))
    except Exception:
        return (0.0, 0.0)


def get_measured_length(entity) -> Optional[float]:
    if entity.dxftype() != "DIMENSION":
        return None
    try:
        return float(entity.get_measurement())
    except Exception:
        return None


def extract_labels(doc, layers: Optional[list[str]] = None) -> list[BarLabel]:
    msp = doc.modelspace()
    layer_filter = set(l.upper() for l in layers) if layers else None
    labels: list[BarLabel] = []
    for entity in msp:
        if entity.dxftype() not in TEXT_BEARING_TYPES:
            continue
        layer = entity.dxf.layer
        if layer_filter is not None and layer.upper() not in layer_filter:
            continue
        raw = get_entity_text(entity)
        if not raw or not raw.strip():
            continue
        cleaned = clean_dxf_text(raw)
        parsed = parse_bar_label(cleaned)
        x, y = get_entity_position(entity)

        defpoint = None
        anchor_source = None
        if entity.dxftype() == "DIMENSION":
            try:
                defpoint = (float(entity.dxf.defpoint[0]), float(entity.dxf.defpoint[1]))
                anchor_source = "defpoint"
            except Exception:
                defpoint = None
        if defpoint is None and entity.dxftype() in ("TEXT", "MTEXT"):
            defpoint = (round(x, 1), round(y, 1))
            anchor_source = "insertion point"

        label = BarLabel(
            layer=layer,
            entity_type=entity.dxftype(),
            handle=entity.dxf.handle,
            raw_text=cleaned,
            x=round(x, 1),
            y=round(y, 1),
            measured_length=get_measured_length(entity),
            defpoint=defpoint,
            anchor_source=anchor_source,
        )
        if parsed.get("parsed_ok"):
            label.parsed_ok = True
            label.count = parsed["count"]
            label.bar_type = parsed["bar_type"]
            label.diameter = parsed["diameter"]
            label.mark = parsed["mark"]
            label.spacing = parsed["spacing"]
            label.suffix = parsed["suffix"]
            label.is_link = parsed["is_link"]  # <-- FIXED: set is_link from parser
        labels.append(label)
    return labels


# ---------------------------------------------------------------------------
# 5. DOT / BAR GEOMETRY
# ---------------------------------------------------------------------------
def find_solid_dots(doc, layer_aliases: tuple[str, ...] = DOT_LAYER_ALIASES) -> list[tuple]:
    msp = doc.modelspace()
    dots = []
    for h in msp.query("HATCH"):
        if not _layer_matches(h.dxf.layer, layer_aliases):
            continue
        if not h.dxf.solid_fill:
            continue
        for path in h.paths:
            for edge in getattr(path, "edges", []):
                center = getattr(edge, "center", None)
                if center is not None:
                    dots.append((float(center[0]), float(center[1])))
            vertices = getattr(path, "vertices", None)
            if vertices and len(vertices) == 2:
                (x1, y1, b1), (x2, y2, b2) = ((v[0], v[1], v[-1]) for v in vertices)
                if abs(abs(b1) - 1.0) <= 0.05 and abs(abs(b2) - 1.0) <= 0.05:
                    dots.append(((x1 + x2) / 2.0, (y1 + y2) / 2.0))
    return dots


def find_donut_dots(doc, layer: Optional[str] = None) -> list[tuple]:
    msp = doc.modelspace()
    query = "LWPOLYLINE" if layer is None else f'LWPOLYLINE[layer=="{layer}"]'
    dots = []
    for p in msp.query(query):
        if not p.closed:
            continue
        points = list(p.get_points())
        if len(points) != 2:
            continue
        (x1, y1, _, _, b1), (x2, y2, _, _, b2) = points
        if abs(abs(b1) - 1.0) > 0.05 or abs(abs(b2) - 1.0) > 0.05:
            continue
        radius = math.hypot(x2 - x1, y2 - y1) / 2
        if not (MIN_DOT_RADIUS_MM <= radius <= MAX_DOT_RADIUS_MM):
            continue
        dots.append(((x1 + x2) / 2, (y1 + y2) / 2))
    return dots


def find_all_dots(doc, hatch_layer_aliases: tuple[str, ...] = DOT_LAYER_ALIASES,
                  donut_layer: Optional[str] = None) -> list[tuple]:
    dots = find_solid_dots(doc, layer_aliases=hatch_layer_aliases) + find_donut_dots(doc, layer=donut_layer)
    deduped = []
    for d in dots:
        if not any(_dist(d, existing) < 1.0 for existing in deduped):
            deduped.append(d)
    return deduped


def find_bar_lines(doc, layer_aliases: tuple[str, ...] = BAR_LAYER_ALIASES) -> list[tuple]:
    msp = doc.modelspace()
    lines = []
    for l in msp.query("LINE"):
        if not _layer_matches(l.dxf.layer, layer_aliases):
            continue
        s = (float(l.dxf.start[0]), float(l.dxf.start[1]))
        e = (float(l.dxf.end[0]), float(l.dxf.end[1]))
        lines.append((s, e, _dist(s, e)))
    return lines


def find_leader_lines(doc, layer_aliases: tuple[str, ...] = LEADER_LINE_LAYER_ALIASES) -> list[tuple]:
    msp = doc.modelspace()
    lines = []
    for l in msp.query("LINE"):
        if not _layer_matches(l.dxf.layer, layer_aliases):
            continue
        s = (float(l.dxf.start[0]), float(l.dxf.start[1]))
        e = (float(l.dxf.end[0]), float(l.dxf.end[1]))
        lines.append((s, e))
    return lines


def _ray_cast_label(near: tuple, point: tuple, label_points: list) -> Optional[tuple]:
    if not label_points:
        return None
    dx, dy = point[0] - near[0], point[1] - near[1]
    length = math.hypot(dx, dy)
    if length == 0:
        return None
    ux, uy = dx / length, dy / length
    hits = []
    for pt, label in label_points:
        vx, vy = pt[0] - point[0], pt[1] - point[1]
        along = vx * ux + vy * uy
        perp = abs(vx * uy - vy * ux)
        if along > 0 and perp <= LEADER_RAY_CORRIDOR_MM:
            hits.append((along, label))
    if not hits:
        return None
    hits.sort(key=lambda h: h[0])
    if len(hits) > 1 and (hits[1][0] - hits[0][0]) < LEADER_RAY_AMBIGUITY_GAP_MM:
        return None
    return (hits[0][1], round(hits[0][0], 1))


def walk_leader_chain(dot: tuple, leader_lines: list, label_points: list) -> Optional[tuple]:
    candidates = []
    for s, e in leader_lines:
        for near, far in ((s, e), (e, s)):
            if abs(near[0] - dot[0]) <= LEADER_X_ALIGN_TOL_MM or abs(near[1] - dot[1]) <= LEADER_X_ALIGN_TOL_MM:
                candidates.append((_dist(dot, near), near, far))
    if not candidates:
        return None
    candidates.sort(key=lambda c: c[0])
    if len(candidates) > 1 and (candidates[1][0] - candidates[0][0]) < LEADER_START_AMBIGUITY_GAP_MM:
        return None
    _, near, point = candidates[0]
    visited = {near}
    for _hop in range(LEADER_MAX_HOPS):
        if label_points:
            nearest_pt, nearest_label = min(label_points, key=lambda lp: _dist(point, lp[0]))
            tip_dist = _dist(point, nearest_pt)
            if tip_dist <= LEADER_TIP_TOL_MM:
                return (nearest_label, round(tip_dist, 1))
        touching = [
            b for s, e in leader_lines for a, b in ((s, e), (e, s))
            if _dist(a, point) <= LEADER_BEND_TOL_MM and b not in visited and _dist(b, point) > LEADER_BEND_TOL_MM
        ]
        if not touching:
            return _ray_cast_label(near, point, label_points)
        if len(touching) > 1:
            return None
        visited.add(point)
        near, point = point, touching[0]
    return _ray_cast_label(near, point, label_points)


def detect_suspicious_length_clusters(bar_lines: list, min_repeat_count: int = 15) -> set:
    counts: dict = defaultdict(int)
    for _, _, length in bar_lines:
        counts[round(length * 2) / 2] += 1
    if not counts:
        return set()
    freqs = sorted(counts.values())
    median_freq = freqs[len(freqs) // 2]
    return {length for length, cnt in counts.items() if cnt >= min_repeat_count and cnt >= median_freq * 5}


def match_dot_lengths(
    labels: list[BarLabel],
    dots: list[tuple],
    bar_lines: list[tuple],
    leader_lines: Optional[list] = None,
    chains: Optional[list] = None,
    dot_label_tol: float = DOT_LABEL_TOL_MM,
    bar_match_tol: float = BAR_MATCH_TOL_MM,
) -> tuple[dict, list[dict]]:
    all_candidates = [l for l in labels if l.parsed_ok and l.defpoint is not None]
    label_points = [(l.defpoint, l) for l in all_candidates]
    leader_lines = leader_lines or []
    suspicious_lengths = detect_suspicious_length_clusters(bar_lines)
    lengths_by_mark: dict = defaultdict(list)
    match_records = []

    for dot in dots:
        record = {"Dot X": round(dot[0], 1), "Dot Y": round(dot[1], 1)}
        match_method = None
        ambiguous = False
        best_label, label_dist = None, None

        if all_candidates:
            aligned = [l for l in all_candidates if _axis_aligned_dist(dot, l.defpoint) <= dot_label_tol]
            if aligned:
                best_label = min(aligned, key=lambda l: _dist(dot, l.defpoint))
                label_dist = _axis_aligned_dist(dot, best_label.defpoint)
                match_method = "axis-aligned"
            else:
                chain_result = walk_leader_chain(dot, leader_lines, label_points)
                if chain_result:
                    best_label, label_dist = chain_result
                    match_method = "leader-chain"
                else:
                    by_dist = sorted(all_candidates, key=lambda l: _dist(dot, l.defpoint))
                    nearest = by_dist[0]
                    nearest_dist = _dist(dot, nearest.defpoint)
                    runner_up_dist = _dist(dot, by_dist[1].defpoint) if len(by_dist) > 1 else float("inf")
                    if nearest_dist <= PROXIMITY_MATCH_MAX_MM and (runner_up_dist - nearest_dist) >= PROXIMITY_AMBIGUITY_GAP_MM:
                        best_label = nearest
                        label_dist = nearest_dist
                        match_method = "nearest-proximity"
                    elif nearest_dist <= PROXIMITY_MATCH_MAX_MM:
                        ambiguous = True

        # <-- FIXED: use label.is_link directly instead of suggest_shape_code()
        is_link = best_label is not None and best_label.is_link
        if bar_lines:
            best_bar_dist, best_bar_len = min(
                ((_point_to_segment_dist(dot, s, e), length) for s, e, length in bar_lines),
                key=lambda x: x[0],
            )
        else:
            best_bar_dist, best_bar_len = None, None

        # Chain-based geometry (for shape-code detection) -- separate from
        # the single-nearest-LINE length match above, since a bent bar's
        # true length/shape needs the whole connected chain, not just
        # whichever individual segment the dot happens to sit closest to.
        best_chain = None
        if chains:
            best_chain_dist, best_chain = min(
                ((_chain_point_dist(dot, c), c) for c in chains),
                key=lambda x: x[0],
            )
        else:
            best_chain_dist = None
        auto_shape_code, shape_candidates = (None, [])
        if best_chain is not None and best_chain_dist is not None and best_chain_dist <= bar_match_tol:
            auto_shape_code, shape_candidates = suggest_shape_from_geometry(
                best_chain.num_segments, best_chain.turn_angles
            )
        record["Num Segments"] = best_chain.num_segments if best_chain else None
        record["Turn Angles"] = str(best_chain.turn_angles) if best_chain else None
        record["Auto Shape Code"] = auto_shape_code
        record["Shape Candidates"] = ",".join(shape_candidates) if shape_candidates else None

        record["Matched Mark"] = best_label.mark if best_label else None
        record["Is Link"] = is_link
        record["Match Method"] = match_method if best_label else ("ambiguous -- not matched" if ambiguous else None)
        record["Label Distance (mm)"] = round(label_dist, 1) if label_dist is not None else None
        record["Bar Line Distance (mm)"] = round(best_bar_dist, 1) if best_bar_dist is not None else None
        record["Bar Length (mm)"] = round(best_bar_len, 1) if best_bar_len is not None else None
        is_suspicious_length = best_bar_len is not None and any(abs(best_bar_len - s) <= 1 for s in suspicious_lengths)
        record["Suspicious Length"] = is_suspicious_length

        if match_method in ("axis-aligned", "leader-chain"):
            label_ok = best_label is not None and not is_link
        elif match_method == "nearest-proximity":
            label_ok = best_label is not None and not is_link
        else:
            label_ok = False
        bar_ok = best_bar_len is not None and best_bar_dist <= bar_match_tol
        record["Accepted"] = bool(label_ok and bar_ok)

        if label_ok and bar_ok:
            lengths_by_mark[best_label.mark].append(
                (round(best_bar_len, 1), match_method, is_suspicious_length, auto_shape_code, shape_candidates)
            )
        match_records.append(record)

    return dict(lengths_by_mark), match_records


# ---------------------------------------------------------------------------
# 6. REPORTING / AGGREGATION
# ---------------------------------------------------------------------------
def build_extraction_report(labels: list[BarLabel]) -> pd.DataFrame:
    rows = []
    for l in labels:
        rows.append({
            "Layer": l.layer,
            "Entity": l.entity_type,
            "Handle": l.handle,
            "Raw Label": l.raw_text,
            "Parsed OK": l.parsed_ok,
            "Count": l.count,
            "Type": l.bar_type,
            "Diameter (mm)": l.diameter,
            "Bar Mark": l.mark,
            "Spacing (mm c/c)": l.spacing,
            "Is Link": l.is_link,  # <-- FIXED: added Is Link column
            "Notes": l.suffix,
            "DXF Measured Length (mm)": round(l.measured_length, 1) if l.measured_length else None,
            "Anchor Source": l.anchor_source,
            "X": l.x,
            "Y": l.y,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Parsed OK", "Bar Mark"], ascending=[False, True], na_position="last")
    return df


def build_summary(labels: list[BarLabel], dot_lengths: Optional[dict] = None) -> pd.DataFrame:
    dot_lengths = dot_lengths or {}
    groups: dict[str, list[BarLabel]] = defaultdict(list)
    for l in labels:
        if not l.parsed_ok:
            continue
        groups[l.mark].append(l)

    rows = []
    conflicts = []
    for mark, items in groups.items():
        types = {i.bar_type for i in items}
        dias = {i.diameter for i in items}
        spacings = {i.spacing for i in items if i.spacing is not None}
        suffixes = sorted({i.suffix for i in items if i.suffix})
        total_count = sum(i.count for i in items)
        conflict = None
        if len(types) > 1 or len(dias) > 1:
            conflict = f"CONFLICT: mark {mark} used with type(s) {types}, diameter(s) {dias} -- check drawing"
            conflicts.append(conflict)

        # <-- FIXED: check if ANY occurrence is a link, not joined suffixes
        is_link = any(i.is_link for i in items)

        if is_link:
            longest_length = None
            valid_lengths = []
            proximity_lengths = []
            suspicious_length_values = []
            auto_shape_code = None
            display_shape_code = None
        else:
            raw_matches = dot_lengths.get(mark, [])
            trusted_matches = [x for x in raw_matches if x[1] in ("axis-aligned", "leader-chain")]
            valid_lengths = sorted(x[0] for x in trusted_matches)
            proximity_lengths = sorted(x[0] for x in raw_matches if x[1] == "nearest-proximity")
            suspicious_length_values = sorted({x[0] for x in trusted_matches if x[2]})
            longest_length = max(valid_lengths) if valid_lengths else None

            # Shape code: only trust an auto-detected code if every trusted
            # geometric instance of this mark agrees on the same single
            # confirmed code. Disagreement (e.g. one instance reads as a
            # straight bar and another as a bent one) is a real signal
            # something's off, not something to average away.
            per_instance_codes = {x[3] for x in trusted_matches if x[3] is not None}
            auto_shape_code = next(iter(per_instance_codes)) if len(per_instance_codes) == 1 else None
            if len(per_instance_codes) > 1:
                shape_conflict = (
                    f"SHAPE MISMATCH: mark {mark}'s traced instances suggest different shape "
                    f"codes ({', '.join(sorted(per_instance_codes))}) -- geometry isn't consistent "
                    f"across occurrences, verify manually"
                )
                conflict = (conflict + "; " + shape_conflict) if conflict else shape_conflict
                conflicts.append(shape_conflict)

            # <-- NEW: if geometry didn't narrow to exactly one confirmed
            # code, don't just go silent -- show what it DID narrow down
            # to (the intersection of every trusted instance's candidate
            # list, so a mark with 3 occurrences that all narrowed to the
            # same 2 candidates shows those 2, not a blank). Still flagged
            # as unconfirmed -- this is "here's what to check", not a
            # length-bearing number being asserted.
            display_shape_code = auto_shape_code
            if auto_shape_code is None and len(per_instance_codes) <= 1:
                candidate_sets = [
                    set(x[4]) - {"99"} for x in trusted_matches if x[4]
                ]
                if candidate_sets:
                    common = set.intersection(*candidate_sets)
                    if common:
                        display_shape_code = "/".join(sorted(common)) + " (UNCONFIRMED -- verify)"

            if len(valid_lengths) > 1 and (valid_lengths[-1] - valid_lengths[0]) > LENGTH_VARIANCE_FLAG_MM:
                spread_note = (
                    f"LENGTH VARIES: mark {mark} has trusted lengths from {valid_lengths[0]} to "
                    f"{valid_lengths[-1]}mm (spread {valid_lengths[-1]-valid_lengths[0]:.1f}mm across "
                    f"{len(valid_lengths)} verified instance(s)) -- confirm with engineer whether this "
                    f"is genuine footing-to-footing variation or a mismatched length"
                )
                conflict = (conflict + "; " + spread_note) if conflict else spread_note
                conflicts.append(spread_note)
            if suspicious_length_values:
                suspicious_note = (
                    f"POSSIBLE DECOY LENGTH: mark {mark} includes length(s) "
                    f"{', '.join(str(v) for v in suspicious_length_values)}mm that repeat unusually often "
                    f"across the drawing (see extraction notes) -- verify before trusting"
                )
                conflict = (conflict + "; " + suspicious_note) if conflict else suspicious_note
                conflicts.append(suspicious_note)

            # <-- NEW: cross-check the label's stated count against how many
            # DISTINCT geometric instances were actually traced for this
            # mark. These currently come from two completely independent
            # sources (label text vs dot/chain matching) and nothing
            # compared them before now -- that's the mark-17 bug: 4 dots
            # tracing correctly to 1 label never surfaced as a count
            # problem because "Total No. Off" only ever read the label text.
            distinct_instances = len(valid_lengths)
            if distinct_instances > 0 and distinct_instances != total_count:
                count_note = (
                    f"COUNT MISMATCH: mark {mark} label text states {total_count} bar(s), but "
                    f"{distinct_instances} distinct geometrically-traced instance(s) were found on "
                    f"the drawing -- verify the true count (this does NOT auto-correct Total No. Off)"
                )
                conflict = (conflict + "; " + count_note) if conflict else count_note
                conflicts.append(count_note)

        rows.append({
            "Bar Mark": mark,
            "Type": next(iter(types)) if len(types) == 1 else "/".join(sorted(types)),
            "Diameter (mm)": next(iter(dias)) if len(dias) == 1 else "/".join(str(d) for d in sorted(dias)),
            "Occurrences on Drawing": len(items),
            "Total No. Off": total_count,
            "Distinct DXF Instances Traced": (len(valid_lengths) if not is_link else None),
            "Spacing (mm c/c)": next(iter(spacings)) if len(spacings) == 1 else ("/".join(str(s) for s in sorted(spacings)) if spacings else None),
            "Notes": "; ".join(suffixes),
            "Is Link": is_link,
            "Suggested Shape Code (verify)": ("LINK" if is_link else display_shape_code),
            "Longest DXF Length (mm)": longest_length,
            "All DXF Lengths (mm)": ", ".join(str(x) for x in valid_lengths) if valid_lengths else None,
            "Length Confidence": ("geometrically verified (axis-aligned/leader-chain)" if valid_lengths else None),
            "Proximity-Matched Lengths (NOT used -- unverified)": ", ".join(str(x) for x in proximity_lengths) if proximity_lengths else None,
            "Flag": conflict or "",
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["_sort"] = df["Bar Mark"].apply(lambda m: (0, int(m)) if str(m).isdigit() else (1, str(m)))
        df = df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    return df


def find_missing_marks(summary_df: pd.DataFrame) -> list[int]:
    if summary_df.empty:
        return []
    non_link = summary_df[~summary_df["Is Link"]]
    numeric_marks = sorted(int(m) for m in non_link["Bar Mark"] if str(m).isdigit())
    if len(numeric_marks) < 2:
        return []
    full_range = set(range(numeric_marks[0], numeric_marks[-1] + 1))
    missing = sorted(full_range - set(numeric_marks))
    return missing


# ---------------------------------------------------------------------------
# 7. WORKBOOK GENERATION
# ---------------------------------------------------------------------------
def _copy_header_block(src_ws: Worksheet, dst_ws: Worksheet, rows: int = 11):
    for row in src_ws.iter_rows(min_row=1, max_row=rows):
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_row <= rows:
            dst_ws.merge_cells(str(merged_range))
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width


def _write_bend_allowance_table(ws: Worksheet, start_row: int = 1) -> int:
    headers = ["Key", "Type", "Dia (mm)", "h (mm)", "n (mm)", "r (mm)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=REF_TABLE_START_COL + j, value=h)
        c.font = Font(name="Arial", bold=True, size=8, italic=True)
    row = start_row + 1
    if shape_codes is not None:
        for steel_type, table in shape_codes.STANDARD_BEND_ALLOWANCES.items():
            for dia, vals in table.items():
                key = f"{steel_type}{dia}"
                values = [key, steel_type, dia, vals["h"], vals["n"], vals["r"]]
                for j, v in enumerate(values):
                    ws.cell(row=row, column=REF_TABLE_START_COL + j, value=v).font = Font(name="Arial", size=8)
                row += 1
    last_row = row - 1
    note_col = get_column_letter(REF_TABLE_START_COL)
    ws.cell(row=start_row - 1 if start_row > 1 else start_row,
            column=REF_TABLE_START_COL,
            value="Standard bend allowances (SANS 282:2011 Fig 2/3) -- used to auto-look-up r for shape codes 37/38/45. "
                  "If the drawing calls for a NON-STANDARD radius (shape code suffix 'S'), this lookup is wrong for that "
                  "row -- check manually.").font = Font(italic=True, size=8)
    return last_row


def _r_lookup_formula(row: int, ref_last_row: int) -> str:
    key_col = get_column_letter(REF_TABLE_START_COL)
    r_col = get_column_letter(REF_TABLE_START_COL + 5)
    return (
        f'IFERROR(INDEX(${r_col}$2:${r_col}${ref_last_row},'
        f'MATCH(C{row}&D{row},${key_col}$2:${key_col}${ref_last_row},0)),"")'
    )


def _shape_length_formula(row: int, ref_last_row: int) -> str:
    r = _r_lookup_formula(row, ref_last_row)
    return (
        f'=IFERROR('
        f'IF(I{row}=20,J{row},'
        f'IF(I{row}=41,J{row}+K{row}+L{row},'
        f'IF(I{row}=42,J{row}+K{row}+L{row}+N{row},'
        f'IF(I{row}=37,J{row}+K{row}-({r})/2-D{row},'
        f'IF(I{row}=38,J{row}+K{row}+L{row}-({r})-2*D{row},'
        f'IF(I{row}=45,J{row}+K{row}+L{row}-({r})/2-D{row},'
        f'IF(I{row}=39,J{row}+0.57*K{row}+L{row}-1.57*D{row},'
        f'"")))))))'
        f',"")'
    )


def write_bbs_workbook(
    summary_df: pd.DataFrame,
    out_path: str,
    template_path: Optional[str] = None,
    member_name: str = "MEMBER",
    item_description: str = "EXTRACTED FROM DXF -- VERIFY LENGTH / SHAPE CODE / BEND DIMENSIONS",
    missing_marks: Optional[list[int]] = None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "BBS from DXF"
    if template_path and Path(template_path).exists():
        src_wb = load_workbook(template_path)
        src_ws = src_wb.active
        _copy_header_block(src_ws, ws, rows=11)
        ws["B8"] = item_description
    else:
        headers = ["Member", "Mark", "Type", "", "No.\nMbrs", "No.\neach", "Total\nNo. off",
                   "Length\nmm", "Shape\nCode", "A\nmm", "B\nmm", "C\nmm",
                   "D\nmm", "n\nmm", "TOTAL WEIGHT\n(TON)"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=10, column=col, value=h)
            c.font = HEADER_FONT
    start_row = 12
    if missing_marks:
        banner_text = (
            f"GAP DETECTED IN BAR MARK NUMBERING: mark(s) {', '.join(str(m) for m in missing_marks)} "
            f"do not appear anywhere on the drawing but fall between marks that do -- "
            f"check whether they were omitted by mistake before relying on this schedule."
        )
        ws.cell(row=start_row, column=1, value=banner_text)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=15)
        banner_cell = ws.cell(row=start_row, column=1)
        banner_cell.font = Font(name="Arial", bold=True, size=9, color="9C0006")
        banner_cell.fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
        banner_cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[start_row].height = 30
        start_row += 1
    first_data_row = start_row
    ref_last_row = _write_bend_allowance_table(ws, start_row=1)
    if shape_codes is None:
        print("[!] sans282_shape_codes.py not found -- shape-code length formulas will be blank.")

    for offset, rec in enumerate(summary_df.to_dict("records")):
        r = start_row + offset
        ws.cell(row=r, column=1, value=member_name if offset == 0 else None)
        ws.cell(row=r, column=2, value=rec["Bar Mark"])
        ws.cell(row=r, column=3, value=rec["Type"])
        ws.cell(row=r, column=4, value=rec["Diameter (mm)"] if isinstance(rec["Diameter (mm)"], int) else None)
        ws.cell(row=r, column=5, value=1)
        ws.cell(row=r, column=6, value=rec["Total No. Off"])
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}")
        longest_length = rec.get("Longest DXF Length (mm)")
        length_formula = _shape_length_formula(r, ref_last_row)
        is_link = bool(rec.get("Is Link"))

        if is_link:
            # Shape 60 is confirmed as this project's real closed-link
            # formula (2A+2B+2n-1.5r-3d) -- but its A/B leg dimensions
            # still aren't traced from closed-loop DXF geometry, so writing
            # "60" here would still be asserting a length-bearing shape
            # code with no dimensions behind it. Flag it as text instead.
            ws.cell(row=r, column=8, value=None)
            ws.cell(row=r, column=8).fill = REVIEW_FILL
            ws.cell(row=r, column=9, value="LINK")
            ws.cell(row=r, column=9).fill = LINK_SHAPE_FILL
            for col in (10, 11):  # A, B dims for link
                ws.cell(row=r, column=col).fill = LINK_SHAPE_FILL
            for col in (12, 13, 14):
                ws.cell(row=r, column=col).fill = REVIEW_FILL
        else:
            if longest_length is not None and str(longest_length).lower() != "nan":
                ws.cell(row=r, column=8, value=float(longest_length))
                ws.cell(row=r, column=8).fill = DXF_LENGTH_FILL
            else:
                ws.cell(row=r, column=8, value=length_formula)
                ws.cell(row=r, column=8).fill = REVIEW_FILL
            # Shape code, in priority order:
            #   1. Exactly one confirmed code from traced geometry -> write
            #      it, green (trustworthy).
            #   2. Geometry narrowed to a shortlist but didn't land on one
            #      -> write the candidates joined with "/", yellow
            #      (something to check, not a blank you have to chase down).
            #   3. Nothing usable at all (unmatched dot, no chain) -> blank,
            #      yellow, for manual entry.
            # Note: option 2 is never a bare number -- it's always the
            # literal candidate text (e.g. "38/54 (UNCONFIRMED -- verify)"),
            # so it can never be mistaken for an asserted, trusted code.
            shape_cell = rec.get("Suggested Shape Code (verify)")
            if shape_cell:
                ws.cell(row=r, column=9, value=shape_cell)
                ws.cell(row=r, column=9).fill = (
                    REVIEW_FILL if "UNCONFIRMED" in str(shape_cell) else DXF_LENGTH_FILL
                )
            else:
                ws.cell(row=r, column=9, value=None)
                ws.cell(row=r, column=9).fill = REVIEW_FILL
            for col in (10, 11, 12, 13, 14):
                ws.cell(row=r, column=col).fill = REVIEW_FILL

        weight_formula = (
            f"=IF(D{r}=8,G{r}*H{r}*0.395,IF(D{r}=10,G{r}*H{r}*0.616,"
            f"IF(D{r}=12,G{r}*H{r}*0.888,IF(D{r}=16,G{r}*H{r}*1.579,"
            f"IF(D{r}=20,G{r}*H{r}*2.466,IF(D{r}=25,G{r}*H{r}*3.854,"
            f"IF(D{r}=32,G{r}*H{r}*6.313))))))/1000/1000"
        )
        ws.cell(row=r, column=15, value=weight_formula)

        def _is_set(v):
            return v is not None and str(v) != "" and str(v).lower() != "nan"
        ref_bits = []
        if _is_set(rec.get("Spacing (mm c/c)")):
            ref_bits.append(f"spacing {rec['Spacing (mm c/c)']} c/c")
        if _is_set(rec.get("Notes")):
            ref_bits.append(str(rec["Notes"]))
        if is_link:
            ref_bits.append("LINK -- shape 60 formula confirmed (2A+2B+2n-1.5r-3d); A/B leg dims need manual entry from drawing")
        else:
            shape_cell = rec.get("Suggested Shape Code (verify)")
            if not shape_cell:
                ref_bits.append("shape code NOT detected (ambiguous or unmatched geometry) -- enter manually")
            elif "UNCONFIRMED" in str(shape_cell):
                ref_bits.append(f"shape code narrowed to candidates {shape_cell} from traced geometry -- pick the right one and confirm")
            else:
                ref_bits.append(f"shape code {shape_cell} auto-detected from traced bar geometry -- verify before use")
        if _is_set(rec.get("All DXF Lengths (mm)")):
            ref_bits.append(f"all DXF lengths found: {rec['All DXF Lengths (mm)']} mm (longest used in column H)")
        if _is_set(rec.get("Proximity-Matched Lengths (NOT used -- unverified)")):
            ref_bits.append(f"NOT USED -- proximity-matched candidate length(s): {rec['Proximity-Matched Lengths (NOT used -- unverified)']} mm")
        if _is_set(rec.get("Flag")):
            ref_bits.append(str(rec["Flag"]))
        ws.cell(row=r, column=17, value="; ".join(ref_bits))
        for col in range(1, 16):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).border = BOX

    last_row = start_row + len(summary_df) - 1
    total_row = last_row + 2
    ws.cell(row=total_row, column=9, value="TOTAL WEIGHT (TON)")
    ws.cell(row=total_row, column=15, value=f"=SUM(O{first_data_row}:O{last_row})")
    ws.cell(row=total_row, column=9).font = HEADER_FONT
    ws.cell(row=total_row, column=15).font = HEADER_FONT
    ws.cell(row=first_data_row - 1, column=17, value="Reference notes (spacing / drawing notes / DXF length -- not part of BBS format)").font = Font(italic=True, size=8)
    ws.column_dimensions["Q"].width = 55
    wb.save(out_path)


# ---------------------------------------------------------------------------
# 8. MAIN PIPELINE
# ---------------------------------------------------------------------------
def process_dxf(
    dxf_path: str,
    template_path: Optional[str] = None,
    member_name: str = "MEMBER",
    layers: Optional[list[str]] = None,
    outdir: Optional[str] = None,
) -> dict:
    """
    Process a DXF file and return all results + file paths.

    Returns:
        {
            "report_df": pd.DataFrame,
            "summary_df": pd.DataFrame,
            "dot_matches_df": pd.DataFrame,
            "bbs_path": str,
            "report_path": str,
            "summary_path": str,
            "dot_match_path": str,
            "logs": list[str],
            "stats": dict,
        }
    """
    logs = []
    stats = {}

    if outdir is None:
        outdir = tempfile.mkdtemp(prefix="bbs_")
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    logs.append(f"Reading {dxf_path} ...")
    doc = ezdxf.readfile(dxf_path)
    labels = extract_labels(doc, layers=layers)
    logs.append(f"Found {len(labels)} text-bearing entities on the target layer(s).")

    parsed = [l for l in labels if l.parsed_ok]
    unparsed = [l for l in labels if not l.parsed_ok]
    logs.append(f"  {len(parsed)} matched the bar-label pattern.")
    if unparsed:
        logs.append(f"  {len(unparsed)} did not match -- listed in extraction_report.csv for manual review.")

    hatch_dots = find_solid_dots(doc, layer_aliases=DOT_LAYER_ALIASES)
    donut_dots = find_donut_dots(doc, layer=None)
    dots = find_all_dots(doc, hatch_layer_aliases=DOT_LAYER_ALIASES, donut_layer=None)
    bar_lines = find_bar_lines(doc, layer_aliases=BAR_LAYER_ALIASES)
    leader_lines = find_leader_lines(doc, layer_aliases=LEADER_LINE_LAYER_ALIASES)
    bar_chains = build_bar_chains(doc, layer_aliases=BAR_LAYER_ALIASES)
    dot_lengths, dot_match_records = match_dot_lengths(
        labels, dots, bar_lines, leader_lines=leader_lines, chains=bar_chains
    )
    accepted = sum(1 for r in dot_match_records if r["Accepted"])
    logs.append(
        f" Leader-dot geometry: {len(hatch_dots)} HATCH-style dot(s) + "
        f"{len(donut_dots)} DONUT-style dot(s) = {len(dots)} total, "
        f"{len(bar_lines)} candidate bar line(s), "
        f"{accepted}/{len(dots)} dot(s) matched to a mark + bar length within tolerance."
    )

    method_counts = Counter(r["Match Method"] for r in dot_match_records if r["Accepted"])
    if method_counts:
        breakdown = ", ".join(f"{v} {k}" for k, v in method_counts.most_common())
        logs.append(f" ({breakdown})")

    unmatched = [r for r in dot_match_records if not r["Accepted"]]
    if unmatched:
        logs.append(f" {len(unmatched)} dot(s) could not be confidently matched -- see dot_bar_matches.csv.")

    report_df = build_extraction_report(labels)
    summary_df = build_summary(labels, dot_lengths=dot_lengths)
    missing_marks = find_missing_marks(summary_df)
    if missing_marks:
        logs.append(
            f"[!] Gap in bar mark numbering -- mark(s) {', '.join(str(m) for m in missing_marks)} "
            f"never appear on the drawing but sit between marks that do. "
            f"Check whether these were skipped by mistake."
        )

    report_path = outdir / "extraction_report.csv"
    summary_path = outdir / "bar_mark_summary.csv"
    dot_match_path = outdir / "dot_bar_matches.csv"
    bbs_path = outdir / "BBS_from_DXF.xlsx"

    report_df.to_csv(report_path, index=False)
    summary_df.to_csv(summary_path, index=False)
    pd.DataFrame(dot_match_records).to_csv(dot_match_path, index=False)
    write_bbs_workbook(
        summary_df, str(bbs_path),
        template_path=template_path,
        member_name=member_name,
        missing_marks=missing_marks,
    )

    logs.append(f"\nWrote outputs to {outdir}")

    stats = {
        "total_entities": len(labels),
        "parsed": len(parsed),
        "unparsed": len(unparsed),
        "dots_total": len(dots),
        "dots_matched": accepted,
        "bar_lines": len(bar_lines),
        "marks_found": len(summary_df),
        "conflicts": len([r for r in summary_df.to_dict("records") if r.get("Flag")]),
        "missing_marks": missing_marks,
    }

    return {
        "report_df": report_df,
        "summary_df": summary_df,
        "dot_matches_df": pd.DataFrame(dot_match_records),
        "bbs_path": str(bbs_path),
        "report_path": str(report_path),
        "summary_path": str(summary_path),
        "dot_match_path": str(dot_match_path),
        "logs": logs,
        "stats": stats,
    }
